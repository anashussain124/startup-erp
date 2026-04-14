"""
Reports Router — PDF and Excel export for each module.
"""
import io
from datetime import date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from auth.dependencies import get_current_user, require_role
from models.hcm import Employee, Payroll
from models.finance import Expense, Revenue
from models.procurement import Vendor, PurchaseOrder, InventoryItem
from models.ppm import Project, Task
from models.crm import Customer, Lead, Sale

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

router = APIRouter(prefix="/api/reports", tags=["Reports"])


# ─── Helpers ─────────────────────────────────────────────────────────────────
def _styled_workbook(title: str, headers: list[str], rows: list[list]):
    """Create a styled Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return wb


def _make_pdf(title: str, headers: list[str], rows: list[list]):
    """Create a styled PDF report."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles["Title"]
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generated: {date.today().isoformat()}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Table
    table_data = [headers] + rows
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
    ]))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return buf


# ─── Employee Report ─────────────────────────────────────────────────────────
@router.get("/hcm/excel")
def hcm_excel(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["ID", "Code", "Name", "Email", "Department", "Position", "Salary", "Hire Date", "Active"]
    employees = db.query(Employee).all()
    rows = [
        [e.id, e.employee_code, f"{e.first_name} {e.last_name}", e.email,
         e.department, e.position, e.salary, str(e.hire_date), "Yes" if e.is_active else "No"]
        for e in employees
    ]
    wb = _styled_workbook("Employees", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_report.xlsx"},
    )


@router.get("/hcm/pdf")
def hcm_pdf(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["ID", "Code", "Name", "Department", "Position", "Salary", "Hire Date"]
    employees = db.query(Employee).all()
    rows = [
        [e.id, e.employee_code, f"{e.first_name} {e.last_name}",
         e.department, e.position, f"${e.salary:,.2f}", str(e.hire_date)]
        for e in employees
    ]
    buf = _make_pdf("Employee Report", headers, rows)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=employees_report.pdf"})


# ─── Finance Report ──────────────────────────────────────────────────────────
@router.get("/finance/excel")
def finance_excel(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    # Revenue sheet
    wb = Workbook()
    # Revenue
    revenues = db.query(Revenue).order_by(Revenue.date.desc()).all()
    rev_headers = ["ID", "Source", "Amount", "Date", "Description"]
    rev_rows = [[r.id, r.source, r.amount, str(r.date), r.description or ""] for r in revenues]
    ws_rev = wb.active
    ws_rev.title = "Revenue"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for i, h in enumerate(rev_headers, 1):
        c = ws_rev.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    for ri, row in enumerate(rev_rows, 2):
        for ci, val in enumerate(row, 1):
            ws_rev.cell(row=ri, column=ci, value=val)

    # Expenses
    expenses = db.query(Expense).order_by(Expense.date.desc()).all()
    exp_headers = ["ID", "Category", "Amount", "Date", "Description", "Approved By"]
    exp_rows = [[e.id, e.category, e.amount, str(e.date), e.description or "", e.approved_by or ""] for e in expenses]
    ws_exp = wb.create_sheet("Expenses")
    for i, h in enumerate(exp_headers, 1):
        c = ws_exp.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    for ri, row in enumerate(exp_rows, 2):
        for ci, val in enumerate(row, 1):
            ws_exp.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finance_report.xlsx"},
    )


@router.get("/finance/pdf")
def finance_pdf(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["Type", "Category/Source", "Amount", "Date", "Description"]
    rows = []
    for r in db.query(Revenue).order_by(Revenue.date.desc()).all():
        rows.append(["Revenue", r.source, f"${r.amount:,.2f}", str(r.date), r.description or ""])
    for e in db.query(Expense).order_by(Expense.date.desc()).all():
        rows.append(["Expense", e.category, f"${e.amount:,.2f}", str(e.date), e.description or ""])
    buf = _make_pdf("Finance Report", headers, rows)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=finance_report.pdf"})


# ─── Procurement Report ─────────────────────────────────────────────────────
@router.get("/procurement/excel")
def procurement_excel(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["PO#", "Vendor ID", "Amount", "Status", "Order Date", "Expected Delivery"]
    pos = db.query(PurchaseOrder).order_by(PurchaseOrder.order_date.desc()).all()
    rows = [[p.po_number, p.vendor_id, p.total_amount, p.status,
             str(p.order_date), str(p.expected_delivery or "")] for p in pos]
    wb = _styled_workbook("Purchase Orders", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=procurement_report.xlsx"},
    )


@router.get("/procurement/pdf")
def procurement_pdf(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["PO#", "Vendor ID", "Amount", "Status", "Order Date"]
    pos = db.query(PurchaseOrder).order_by(PurchaseOrder.order_date.desc()).all()
    rows = [[p.po_number, p.vendor_id, f"${p.total_amount:,.2f}", p.status, str(p.order_date)] for p in pos]
    buf = _make_pdf("Procurement Report", headers, rows)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=procurement_report.pdf"})


# ─── PPM Report ──────────────────────────────────────────────────────────────
@router.get("/ppm/excel")
def ppm_excel(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["ID", "Project", "Status", "Priority", "Budget", "Spent", "Start", "End", "Manager"]
    projects = db.query(Project).all()
    rows = [[p.id, p.name, p.status, p.priority, p.budget, p.spent,
             str(p.start_date), str(p.end_date or ""), p.manager or ""] for p in projects]
    wb = _styled_workbook("Projects", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=projects_report.xlsx"},
    )


@router.get("/ppm/pdf")
def ppm_pdf(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["Project", "Status", "Priority", "Budget", "Spent", "Start", "End"]
    projects = db.query(Project).all()
    rows = [[p.name, p.status, p.priority, f"${p.budget:,.2f}",
             f"${p.spent:,.2f}", str(p.start_date), str(p.end_date or "")] for p in projects]
    buf = _make_pdf("Project Portfolio Report", headers, rows)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=projects_report.pdf"})


# ─── CRM Report ──────────────────────────────────────────────────────────────
@router.get("/crm/excel")
def crm_excel(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["ID", "Name", "Email", "Company", "Segment", "Lifetime Value", "Active"]
    customers = db.query(Customer).all()
    rows = [[c.id, c.name, c.email, c.company or "", c.segment or "",
             c.lifetime_value, "Yes" if c.is_active else "No"] for c in customers]
    wb = _styled_workbook("Customers", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers_report.xlsx"},
    )


@router.get("/crm/pdf")
def crm_pdf(db: Session = Depends(get_db), user=Depends(require_role("admin", "manager"))):
    headers = ["Name", "Email", "Company", "Segment", "Lifetime Value"]
    customers = db.query(Customer).all()
    rows = [[c.name, c.email, c.company or "", c.segment or "",
             f"${c.lifetime_value:,.2f}"] for c in customers]
    buf = _make_pdf("CRM Customer Report", headers, rows)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=customers_report.pdf"})
